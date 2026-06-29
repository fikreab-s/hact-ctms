"""
Generate OpenClinica CE CRF Excel Upload Files — PSBI RCT 2
=============================================================
Creates 4 CRF Excel files in OpenClinica's exact upload format:
  1. A1_SCREEN_ENROLL  — Screening & Enrollment
  2. B1_RCT2_48H       — 48-Hour Assessment
  3. B2_RCT2_TREATMENT — Treatment Record
  4. C1_SAE             — Serious Adverse Event

Each file has 4 required sheets: CRF, Sections, Groups, Items

Usage:
  python generate_psbi_crfs.py

Output:
  openclinica/crfs/A1_SCREEN_ENROLL_v1.0.xls
  openclinica/crfs/B1_RCT2_48H_v1.0.xls
  openclinica/crfs/B2_RCT2_TREATMENT_v1.0.xls
  openclinica/crfs/C1_SAE_v1.0.xls
"""

import os
from openpyxl import Workbook

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "crfs")
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ── Reusable Codelists ──────────────────────────────────────────────────
CODELISTS = {
    "YN": {
        "options": "Yes,No",
        "values": "1,2",
    },
    "SEX": {
        "options": "Male,Female,Unknown",
        "values": "M,F,U",
    },
    "ASSESSLOC": {
        "options": "Hospital,Health Centre,Health Post,Home,Other",
        "values": "1,2,3,4,5",
    },
    "RESPONDENT": {
        "options": "Mother,Father,Grandmother,Grandfather,Other relative,Other",
        "values": "1,2,3,4,5,6",
    },
    "AEACN": {
        "options": "Drug withdrawn,Drug interrupted,Dose reduced,Dose not changed,Not applicable,Unknown",
        "values": "1,2,3,4,5,6",
    },
    "AEOUT": {
        "options": "Recovered,Recovering,Not recovered,Fatal,Unknown",
        "values": "1,2,3,4,5",
    },
    "TRTARM": {
        "options": "Continued Inpatient Treatment,Discharged on Oral Amoxicillin",
        "values": "1,2",
    },
    "VISIT": {
        "options": "Day 2,Day 4,Day 8",
        "values": "D2,D4,D8",
    },
    "FEEDSTATUS": {
        "options": "No difficulty,Some difficulty,Not able to feed at all",
        "values": "1,2,3",
    },
}

# ── Item Column Headers (must be EXACT for OpenClinica upload) ──────────
ITEM_HEADERS = [
    "ITEM_NAME",
    "DESCRIPTION_LABEL",
    "LEFT_ITEM_TEXT",
    "UNITS",
    "RIGHT_ITEM_TEXT",
    "SECTION_LABEL",
    "GROUP_LABEL",
    "HEADER",
    "SUBHEADER",
    "PARENT_ITEM",
    "COLUMN_NUMBER",
    "PAGE_NUMBER",
    "QUESTION_NUMBER",
    "RESPONSE_TYPE",
    "RESPONSE_LABEL",
    "RESPONSE_OPTIONS_TEXT",
    "RESPONSE_VALUES_OR_CALCULATIONS",
    "RESPONSE_LAYOUT",
    "DEFAULT_VALUE",
    "DATA_TYPE",
    "WIDTH_DECIMAL",
    "VALIDATION",
    "VALIDATION_ERROR_MESSAGE",
    "PHI",
    "REQUIRED",
    "ITEM_DISPLAY_STATUS",
    "SIMPLE_CONDITIONAL_DISPLAY",
]


def make_item(
    name, label, section, group, response_type="text", data_type="ST",
    codelist=None, units="", required="1", phi="0", validation="",
    validation_msg="", header="", subheader="", question_num="",
    width_decimal="", layout="", display_status="", conditional="",
):
    """Build one item row dict."""
    if codelist and codelist in CODELISTS:
        cl = CODELISTS[codelist]
        options_text = cl["options"]
        values = cl["values"]
        resp_label = codelist.lower()
    else:
        options_text = ""
        values = ""
        resp_label = name.lower() if response_type == "text" else ""

    # For text/date/numeric, response_label = item name
    if response_type in ("text", "textarea"):
        resp_label = name.lower()
        options_text = ""
        values = ""

    return {
        "ITEM_NAME": name,
        "DESCRIPTION_LABEL": label,
        "LEFT_ITEM_TEXT": label,
        "UNITS": units,
        "RIGHT_ITEM_TEXT": "",
        "SECTION_LABEL": section,
        "GROUP_LABEL": group,
        "HEADER": header,
        "SUBHEADER": subheader,
        "PARENT_ITEM": "",
        "COLUMN_NUMBER": "",
        "PAGE_NUMBER": "",
        "QUESTION_NUMBER": question_num,
        "RESPONSE_TYPE": response_type,
        "RESPONSE_LABEL": resp_label,
        "RESPONSE_OPTIONS_TEXT": options_text,
        "RESPONSE_VALUES_OR_CALCULATIONS": values,
        "RESPONSE_LAYOUT": layout if layout else ("Horizontal" if response_type == "radio" else ""),
        "DEFAULT_VALUE": "",
        "DATA_TYPE": data_type,
        "WIDTH_DECIMAL": width_decimal,
        "VALIDATION": validation,
        "VALIDATION_ERROR_MESSAGE": validation_msg,
        "PHI": phi,
        "REQUIRED": required,
        "ITEM_DISPLAY_STATUS": display_status,
        "SIMPLE_CONDITIONAL_DISPLAY": conditional,
    }


def write_crf_excel(filename, crf_name, version, description, sections, groups, items):
    """Write one OpenClinica CRF Excel file with 4 sheets."""
    wb = Workbook()

    # ── Sheet 1: CRF ──
    ws_crf = wb.active
    ws_crf.title = "CRF"
    ws_crf.append(["CRF_NAME", "VERSION", "VERSION_DESCRIPTION", "REVISION_NOTES"])
    ws_crf.append([crf_name, version, description, "Initial version"])

    # ── Sheet 2: Sections ──
    ws_sec = wb.create_sheet("Sections")
    ws_sec.append(["SECTION_LABEL", "SECTION_TITLE", "SUBTITLE", "INSTRUCTIONS", "PAGE_NUMBER", "PARENT_SECTION"])
    for sec in sections:
        ws_sec.append([
            sec["label"], sec["title"], sec.get("subtitle", ""),
            sec.get("instructions", ""), sec.get("page", ""), sec.get("parent", ""),
        ])

    # ── Sheet 3: Groups ──
    ws_grp = wb.create_sheet("Groups")
    ws_grp.append(["GROUP_LABEL", "GROUP_LAYOUT", "GROUP_HEADER", "GROUP_REPEAT_NUMBER", "GROUP_REPEAT_MAX"])
    for grp in groups:
        ws_grp.append([
            grp["label"], grp.get("layout", ""), grp.get("header", ""),
            grp.get("repeat_num", ""), grp.get("repeat_max", ""),
        ])

    # ── Sheet 4: Items ──
    ws_items = wb.create_sheet("Items")
    ws_items.append(ITEM_HEADERS)
    for item in items:
        row = [item.get(h, "") for h in ITEM_HEADERS]
        ws_items.append(row)

    # ── Column widths for readability ──
    for ws in [ws_crf, ws_sec, ws_grp, ws_items]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    print(f"  ✅ Created: {filepath}")
    return filepath


# =====================================================================
# CRF 1: A1_SCREEN_ENROLL — Screening & Enrollment
# =====================================================================
def build_a1_screen_enroll():
    sections = [
        {"label": "screening", "title": "Screening Information",
         "subtitle": "Part A: Subject identification and screening",
         "instructions": "Complete this form at the time of initial screening. Record all required information accurately."},
        {"label": "demographics", "title": "Infant Demographics",
         "subtitle": "Date of birth, sex, and weight",
         "instructions": "Record the infant's demographic data as observed at screening."},
        {"label": "eligibility", "title": "Eligibility Assessment",
         "subtitle": "Inclusion and exclusion criteria",
         "instructions": "Assess all inclusion and exclusion criteria. Both must be satisfied for enrollment."},
        {"label": "enrollment", "title": "Enrollment & Randomization",
         "subtitle": "Treatment arm assignment",
         "instructions": "Complete only if the subject meets eligibility criteria and consent is obtained."},
    ]

    groups = [
        {"label": "screening_group", "header": "Screening Data"},
        {"label": "demographics_group", "header": "Infant Demographics"},
        {"label": "eligibility_group", "header": "Eligibility Criteria"},
        {"label": "enrollment_group", "header": "Enrollment Data"},
    ]

    items = [
        # ── Screening Section ──
        make_item("SUBJID", "Identification number of the young infant",
                   "screening", "screening_group", "text", "ST",
                   header="Subject Identification", question_num="A1.1"),
        make_item("SCRDTC", "Date of screening",
                   "screening", "screening_group", "text", "DATE",
                   question_num="A1.2"),
        make_item("SCRTM", "Time of screening (HH:MM)",
                   "screening", "screening_group", "text", "ST",
                   required="0", question_num="A1.3",
                   validation="regexp: /^([01]\\d|2[0-3]):([0-5]\\d)$/",
                   validation_msg="Enter time in HH:MM format (e.g., 09:30)"),
        make_item("SCRCONS", "Consent for screening obtained from parent/guardian",
                   "screening", "screening_group", "radio", "INT",
                   codelist="YN", question_num="A1.4",
                   header="Consent"),

        # ── Demographics Section ──
        make_item("BRTHDTC", "Date of birth of the infant",
                   "demographics", "demographics_group", "text", "DATE",
                   question_num="A1.5", phi="1",
                   header="Infant Demographics"),
        make_item("AGE", "Age at screening (days)",
                   "demographics", "demographics_group", "text", "INT",
                   required="0", question_num="A1.6", units="days",
                   validation="range(0, 59)",
                   validation_msg="Age must be between 0 and 59 days"),
        make_item("SEX", "Sex of the infant",
                   "demographics", "demographics_group", "radio", "ST",
                   codelist="SEX", question_num="A1.7"),
        make_item("WEIGHT", "Weight of the infant",
                   "demographics", "demographics_group", "text", "REAL",
                   units="grams", question_num="A1.8",
                   validation="range(500, 9000)",
                   validation_msg="Weight must be between 500 and 9000 grams"),

        # ── Eligibility Section ──
        make_item("INCLMET", "All inclusion criteria met",
                   "eligibility", "eligibility_group", "radio", "INT",
                   codelist="YN", question_num="A1.9",
                   header="Inclusion Criteria"),
        make_item("INCL1", "Age 0-59 days",
                   "eligibility", "eligibility_group", "radio", "INT",
                   codelist="YN", question_num="A1.9a",
                   subheader="Individual inclusion criteria"),
        make_item("INCL2", "Signs of possible serious bacterial infection",
                   "eligibility", "eligibility_group", "radio", "INT",
                   codelist="YN", question_num="A1.9b"),
        make_item("INCL3", "Parent/guardian willing to provide informed consent",
                   "eligibility", "eligibility_group", "radio", "INT",
                   codelist="YN", question_num="A1.9c"),
        make_item("EXCLMET", "Any exclusion criteria present",
                   "eligibility", "eligibility_group", "radio", "INT",
                   codelist="YN", question_num="A1.10",
                   header="Exclusion Criteria"),
        make_item("EXCL1", "Weight less than 1500 grams",
                   "eligibility", "eligibility_group", "radio", "INT",
                   codelist="YN", question_num="A1.10a",
                   subheader="Individual exclusion criteria"),
        make_item("EXCL2", "Signs of critical illness requiring intensive care",
                   "eligibility", "eligibility_group", "radio", "INT",
                   codelist="YN", question_num="A1.10b"),
        make_item("EXCL3", "Known congenital malformation incompatible with survival",
                   "eligibility", "eligibility_group", "radio", "INT",
                   codelist="YN", question_num="A1.10c"),

        # ── Enrollment Section ──
        make_item("ENROLLED", "Subject enrolled in the study",
                   "enrollment", "enrollment_group", "radio", "INT",
                   codelist="YN", question_num="A1.11",
                   header="Enrollment Decision"),
        make_item("ENRLDTC", "Date of enrollment",
                   "enrollment", "enrollment_group", "text", "DATE",
                   required="0", question_num="A1.12",
                   conditional="ENROLLED,1"),
        make_item("RNDNUM", "Randomization number",
                   "enrollment", "enrollment_group", "text", "ST",
                   required="0", question_num="A1.13",
                   conditional="ENROLLED,1"),
        make_item("TRTARM", "Treatment arm assigned",
                   "enrollment", "enrollment_group", "single-select", "INT",
                   codelist="TRTARM", required="0", question_num="A1.14",
                   conditional="ENROLLED,1"),
    ]

    return write_crf_excel(
        "A1_SCREEN_ENROLL_v1.0.xlsx",
        "A1_SCREEN_ENROLL", "v1.0",
        "Screening and Enrollment Form for PSBI RCT 2 — captures subject identification, consent, demographics, eligibility criteria, and treatment arm randomization",
        sections, groups, items,
    )


# =====================================================================
# CRF 2: B1_RCT2_48H — 48-Hour Assessment
# =====================================================================
def build_b1_rct2_48h():
    sections = [
        {"label": "assessment", "title": "48-Hour Clinical Assessment",
         "subtitle": "RCT 2: Assessment at 48 hours post-enrollment",
         "instructions": "Complete this form 48 hours (±6 hours) after enrollment. This assessment determines whether the infant can be safely discharged on oral amoxicillin."},
        {"label": "vitals", "title": "Vital Signs",
         "subtitle": "Temperature and clinical observations",
         "instructions": "Measure axillary temperature with a calibrated digital thermometer."},
        {"label": "clinical_signs", "title": "Clinical Signs Assessment",
         "subtitle": "Danger signs and clinical evaluation",
         "instructions": "Assess each clinical sign carefully. Any critical illness sign requires continued inpatient treatment."},
        {"label": "decision", "title": "Clinical Decision",
         "subtitle": "Discharge or continue inpatient treatment",
         "instructions": "Based on the clinical assessment, record the management decision."},
    ]

    groups = [
        {"label": "assess_group", "header": "Assessment Information"},
        {"label": "vitals_group", "header": "Vital Signs"},
        {"label": "signs_group", "header": "Clinical Signs"},
        {"label": "decision_group", "header": "Management Decision"},
    ]

    items = [
        # ── Assessment Section ──
        make_item("SUBJID", "Identification number",
                   "assessment", "assess_group", "text", "ST",
                   question_num="B1.1", header="Subject Identification"),
        make_item("ASSESSDTC", "Date of assessment",
                   "assessment", "assess_group", "text", "DATE",
                   question_num="B1.2"),
        make_item("ASSESSTM", "Time of assessment (HH:MM)",
                   "assessment", "assess_group", "text", "ST",
                   required="0", question_num="B1.3"),
        make_item("ASSESSLOC", "Place of assessment",
                   "assessment", "assess_group", "single-select", "INT",
                   codelist="ASSESSLOC", question_num="B1.4"),

        # ── Vital Signs Section ──
        make_item("TEMP", "Axillary temperature",
                   "vitals", "vitals_group", "text", "REAL",
                   units="°C", question_num="B1.5",
                   header="Temperature",
                   validation="range(34.0, 42.0)",
                   validation_msg="Temperature must be between 34.0 and 42.0 °C"),
        make_item("RESPRATE", "Respiratory rate",
                   "vitals", "vitals_group", "text", "INT",
                   units="breaths/min", required="0", question_num="B1.6",
                   validation="range(10, 100)",
                   validation_msg="Respiratory rate must be between 10 and 100"),
        make_item("HEARTRATE", "Heart rate",
                   "vitals", "vitals_group", "text", "INT",
                   units="beats/min", required="0", question_num="B1.7",
                   validation="range(60, 220)",
                   validation_msg="Heart rate must be between 60 and 220"),

        # ── Clinical Signs Section ──
        make_item("FEEDDIFF", "Feeding difficulty",
                   "clinical_signs", "signs_group", "single-select", "INT",
                   codelist="FEEDSTATUS", question_num="B1.8",
                   header="Feeding and Danger Signs"),
        make_item("CHESTIND", "Severe chest indrawing",
                   "clinical_signs", "signs_group", "radio", "INT",
                   codelist="YN", question_num="B1.9"),
        make_item("CONVULS", "Convulsions observed",
                   "clinical_signs", "signs_group", "radio", "INT",
                   codelist="YN", required="0", question_num="B1.10"),
        make_item("LETHARGY", "Lethargy or unconsciousness",
                   "clinical_signs", "signs_group", "radio", "INT",
                   codelist="YN", required="0", question_num="B1.11"),
        make_item("HYPOTHER", "Hypothermia (<35.5 °C)",
                   "clinical_signs", "signs_group", "radio", "INT",
                   codelist="YN", required="0", question_num="B1.12"),
        make_item("HYPERTHER", "Hyperthermia (≥38.0 °C)",
                   "clinical_signs", "signs_group", "radio", "INT",
                   codelist="YN", required="0", question_num="B1.13"),
        make_item("CRITILL", "Any critical illness signs present",
                   "clinical_signs", "signs_group", "radio", "INT",
                   codelist="YN", question_num="B1.14",
                   header="Overall Assessment"),

        # ── Decision Section ──
        make_item("DSCHRG", "Decision: Infant can be discharged",
                   "decision", "decision_group", "radio", "INT",
                   codelist="YN", question_num="B1.15",
                   header="Management Decision"),
        make_item("DSCHREASON", "If not discharged, reason for continued treatment",
                   "decision", "decision_group", "textarea", "ST",
                   required="0", question_num="B1.16",
                   conditional="DSCHRG,2"),
    ]

    return write_crf_excel(
        "B1_RCT2_48H_v1.0.xlsx",
        "B1_RCT2_48H", "v1.0",
        "48-Hour Assessment Form for PSBI RCT 2 — evaluates clinical signs, vital signs, and determines discharge eligibility",
        sections, groups, items,
    )


# =====================================================================
# CRF 3: B2_RCT2_TREATMENT — Treatment Record
# =====================================================================
def build_b2_rct2_treatment():
    sections = [
        {"label": "visit_info", "title": "Visit Information",
         "subtitle": "Treatment follow-up visit details",
         "instructions": "Complete this form at each treatment follow-up visit (Day 2, Day 4, Day 8)."},
        {"label": "treatment", "title": "Treatment Information",
         "subtitle": "Medication compliance and administration",
         "instructions": "Record current treatment status, compliance, and any changes to the regimen."},
        {"label": "clinical_status", "title": "Clinical Status",
         "subtitle": "Current clinical condition of the infant",
         "instructions": "Assess the infant's current clinical condition."},
    ]

    groups = [
        {"label": "visit_group", "header": "Visit Details"},
        {"label": "treatment_group", "header": "Treatment Details"},
        {"label": "status_group", "header": "Clinical Status"},
    ]

    items = [
        # ── Visit Info Section ──
        make_item("SUBJID", "Identification number",
                   "visit_info", "visit_group", "text", "ST",
                   question_num="B2.1", header="Subject Identification"),
        make_item("DAYNUM", "Visit day number",
                   "visit_info", "visit_group", "single-select", "ST",
                   codelist="VISIT", question_num="B2.2"),
        make_item("VISITDTC", "Date of visit",
                   "visit_info", "visit_group", "text", "DATE",
                   question_num="B2.3"),
        make_item("ASSESSLOC", "Place of completing this form",
                   "visit_info", "visit_group", "single-select", "INT",
                   codelist="ASSESSLOC", question_num="B2.4"),
        make_item("RESPONDENT", "Adult primary respondent present",
                   "visit_info", "visit_group", "single-select", "INT",
                   codelist="RESPONDENT", required="0", question_num="B2.5"),

        # ── Treatment Section ──
        make_item("AMOXGIVEN", "Oral amoxicillin given today",
                   "treatment", "treatment_group", "radio", "INT",
                   codelist="YN", question_num="B2.6",
                   header="Medication Compliance"),
        make_item("REMTAB", "Remaining tablets/suspension (count or mL)",
                   "treatment", "treatment_group", "text", "REAL",
                   required="0", question_num="B2.7",
                   validation="range(0, 100)",
                   validation_msg="Must be between 0 and 100"),
        make_item("MISSEDDOSES", "Number of missed doses since last visit",
                   "treatment", "treatment_group", "text", "INT",
                   required="0", question_num="B2.8",
                   validation="range(0, 20)",
                   validation_msg="Must be between 0 and 20"),
        make_item("OTHABX", "Other antibiotic given (name)",
                   "treatment", "treatment_group", "text", "ST",
                   required="0", question_num="B2.9",
                   header="Additional Medications"),
        make_item("OTHABXREAS", "Reason for other antibiotic",
                   "treatment", "treatment_group", "textarea", "ST",
                   required="0", question_num="B2.10"),

        # ── Clinical Status Section ──
        make_item("FEEDSTAT", "Feeding status at this visit",
                   "clinical_status", "status_group", "single-select", "INT",
                   codelist="FEEDSTATUS", required="0", question_num="B2.11",
                   header="Clinical Assessment"),
        make_item("TEMP", "Axillary temperature",
                   "clinical_status", "status_group", "text", "REAL",
                   units="°C", required="0", question_num="B2.12",
                   validation="range(34.0, 42.0)",
                   validation_msg="Temperature must be between 34.0 and 42.0 °C"),
        make_item("IMPROVING", "Clinical condition improving",
                   "clinical_status", "status_group", "radio", "INT",
                   codelist="YN", required="0", question_num="B2.13"),
        make_item("HOSPREF", "Referred/readmitted to hospital",
                   "clinical_status", "status_group", "radio", "INT",
                   codelist="YN", required="0", question_num="B2.14"),
        make_item("HOSPREFREAS", "If referred, reason for referral",
                   "clinical_status", "status_group", "textarea", "ST",
                   required="0", question_num="B2.15",
                   conditional="HOSPREF,1"),
    ]

    return write_crf_excel(
        "B2_RCT2_TREATMENT_v1.0.xlsx",
        "B2_RCT2_TREATMENT", "v1.0",
        "Treatment Record Form for PSBI RCT 2 — tracks medication compliance, clinical status, and management at Days 2, 4, and 8",
        sections, groups, items,
    )


# =====================================================================
# CRF 4: C1_SAE — Serious Adverse Event
# =====================================================================
def build_c1_sae():
    sections = [
        {"label": "sae_id", "title": "SAE Identification",
         "subtitle": "Subject and report information",
         "instructions": "Complete this form whenever a Serious Adverse Event occurs. Report fatal/life-threatening SAEs within 7 days, all others within 15 days."},
        {"label": "sae_type", "title": "Type of Adverse Event",
         "subtitle": "Classification of the adverse event",
         "instructions": "Check all types of adverse event that apply to this subject."},
        {"label": "sae_dates", "title": "SAE Dates & Causality",
         "subtitle": "Onset, resolution, and relationship to study treatment",
         "instructions": "Record all dates accurately. If the event is ongoing, leave resolution date blank."},
        {"label": "sae_outcome", "title": "Outcome & Action",
         "subtitle": "Action taken and final outcome",
         "instructions": "Record the action taken with study drug and the final outcome of the event."},
        {"label": "sae_narrative", "title": "Narrative Description",
         "subtitle": "Detailed description of the event",
         "instructions": "Provide a detailed clinical narrative of the event including circumstances, treatment given, and current status."},
    ]

    groups = [
        {"label": "sae_id_group", "header": "Identification"},
        {"label": "sae_type_group", "header": "Event Classification"},
        {"label": "sae_dates_group", "header": "Dates & Causality"},
        {"label": "sae_outcome_group", "header": "Outcome"},
        {"label": "sae_narrative_group", "header": "Narrative"},
    ]

    items = [
        # ── SAE Identification ──
        make_item("SUBJID", "Identification number of the young infant",
                   "sae_id", "sae_id_group", "text", "ST",
                   question_num="C1.1", header="Subject Identification"),
        make_item("SAERPTDTC", "Date of filling this SAE form",
                   "sae_id", "sae_id_group", "text", "DATE",
                   question_num="C1.2"),
        make_item("SAERPTNUM", "SAE report number (for this subject)",
                   "sae_id", "sae_id_group", "text", "INT",
                   question_num="C1.3",
                   validation="range(1, 99)",
                   validation_msg="SAE number must be between 1 and 99"),

        # ── SAE Type ──
        make_item("ANAPHYL", "Anaphylactic reaction",
                   "sae_type", "sae_type_group", "radio", "INT",
                   codelist="YN", question_num="C1.4",
                   header="Adverse Event Types (check all that apply)"),
        make_item("ALLERGIC", "Other allergic reaction / rash",
                   "sae_type", "sae_type_group", "radio", "INT",
                   codelist="YN", question_num="C1.5"),
        make_item("INJSITE", "Injection site infection / abscess",
                   "sae_type", "sae_type_group", "radio", "INT",
                   codelist="YN", question_num="C1.6"),
        make_item("DIARRHEA", "Diarrhoea with severe dehydration",
                   "sae_type", "sae_type_group", "radio", "INT",
                   codelist="YN", question_num="C1.7"),
        make_item("AESLIFE", "Life-threatening adverse event",
                   "sae_type", "sae_type_group", "radio", "INT",
                   codelist="YN", question_num="C1.8"),
        make_item("AESDEATH", "Death",
                   "sae_type", "sae_type_group", "radio", "INT",
                   codelist="YN", question_num="C1.9"),
        make_item("AEOTHER", "Other serious adverse event",
                   "sae_type", "sae_type_group", "radio", "INT",
                   codelist="YN", question_num="C1.10"),
        make_item("AEOTHERSP", "If other, specify",
                   "sae_type", "sae_type_group", "text", "ST",
                   required="0", question_num="C1.10a",
                   conditional="AEOTHER,1"),

        # ── SAE Dates & Causality ──
        make_item("AESTDTC", "Date of onset of the adverse event",
                   "sae_dates", "sae_dates_group", "text", "DATE",
                   question_num="C1.11", header="Event Timeline"),
        make_item("AEENDTC", "Date of resolution (leave blank if ongoing)",
                   "sae_dates", "sae_dates_group", "text", "DATE",
                   required="0", question_num="C1.12"),
        make_item("AEREL", "Relationship to study treatment",
                   "sae_dates", "sae_dates_group", "single-select", "INT",
                   question_num="C1.13",
                   header="Causality Assessment",
                   codelist=None),  # custom codelist below

        # ── SAE Outcome ──
        make_item("AEACN", "Action taken with study drug",
                   "sae_outcome", "sae_outcome_group", "single-select", "INT",
                   codelist="AEACN", question_num="C1.14",
                   header="Action & Outcome"),
        make_item("AEOUT", "Outcome of the adverse event",
                   "sae_outcome", "sae_outcome_group", "single-select", "INT",
                   codelist="AEOUT", question_num="C1.15"),
        make_item("DTHDTC", "If fatal — date of death",
                   "sae_outcome", "sae_outcome_group", "text", "DATE",
                   required="0", phi="1", question_num="C1.16",
                   conditional="AESDEATH,1"),
        make_item("DTHCAUSE", "If fatal — primary cause of death",
                   "sae_outcome", "sae_outcome_group", "text", "ST",
                   required="0", question_num="C1.17",
                   conditional="AESDEATH,1"),

        # ── SAE Narrative ──
        make_item("SAEDESC", "Detailed description of the SAE / death",
                   "sae_narrative", "sae_narrative_group", "textarea", "ST",
                   question_num="C1.18",
                   header="Clinical Narrative"),
        make_item("SAETRT", "Treatment given for the SAE",
                   "sae_narrative", "sae_narrative_group", "textarea", "ST",
                   required="0", question_num="C1.19"),
        make_item("RPTNAME", "Name of person completing this report",
                   "sae_narrative", "sae_narrative_group", "text", "ST",
                   question_num="C1.20",
                   header="Reporter Information"),
        make_item("RPTDESIG", "Designation / role",
                   "sae_narrative", "sae_narrative_group", "text", "ST",
                   required="0", question_num="C1.21"),
    ]

    # Fix AEREL — custom codelist (not in standard codelists)
    for item in items:
        if item["ITEM_NAME"] == "AEREL":
            item["RESPONSE_LABEL"] = "aerel"
            item["RESPONSE_OPTIONS_TEXT"] = "Not related,Unlikely,Possible,Probable,Definite"
            item["RESPONSE_VALUES_OR_CALCULATIONS"] = "1,2,3,4,5"

    return write_crf_excel(
        "C1_SAE_v1.0.xlsx",
        "C1_SAE", "v1.0",
        "Serious Adverse Event Form for PSBI study — captures SAE type, timeline, causality, action, outcome, and clinical narrative for CIOMS I reporting",
        sections, groups, items,
    )


# =====================================================================
# Main — Generate all 4 CRF files
# =====================================================================
if __name__ == "__main__":
    print("\n🏥 Generating PSBI RCT 2 CRF Excel Files for OpenClinica CE Upload\n")
    print("=" * 65)

    files = []
    files.append(build_a1_screen_enroll())
    files.append(build_b1_rct2_48h())
    files.append(build_b2_rct2_treatment())
    files.append(build_c1_sae())

    print("\n" + "=" * 65)
    print(f"\n✅ All {len(files)} CRF files generated successfully!")
    print(f"📁 Output directory: {os.path.abspath(OUTPUT_DIR)}")
    print("\n📋 Upload Instructions:")
    print("   1. Login to OpenClinica → Tasks → Build CRF")
    print("   2. Click 'Create a New CRF' → Select 'Upload CRF from Excel'")
    print("   3. Browse to the .xlsx file → Click Upload")
    print("   4. Review the CRF preview → Click 'Submit'")
    print("   5. Repeat for all 4 CRF files")
    print(f"\n   Files to upload (in order):")
    for f in files:
        print(f"     • {os.path.basename(f)}")
    print()
