import sys, os
# Read OpenClinica CRF template and dump all sheet headers
try:
    import xlrd
    wb = xlrd.open_workbook(r'C:\Users\hello\Desktop\HACT project\docs\CRF_Design_Template_v3.9.xls')
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        print(f"\n=== SHEET: {name} ({ws.nrows} rows, {ws.ncols} cols) ===")
        for r in range(min(3, ws.nrows)):
            row = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
            print(f"  Row {r}: {row}")
except ImportError:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(r'C:\Users\hello\Desktop\HACT project\docs\CRF_Design_Template_v3.9.xls')
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"\n=== SHEET: {name} ({ws.max_row} rows, {ws.max_column} cols) ===")
            for r, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
                print(f"  Row {r}: {list(row)}")
    except Exception as e2:
        print(f"openpyxl error: {e2}")
        # Try pandas as fallback
        try:
            import pandas as pd
            xls = pd.ExcelFile(r'C:\Users\hello\Desktop\HACT project\docs\CRF_Design_Template_v3.9.xls')
            for name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=name, header=None, nrows=5)
                print(f"\n=== SHEET: {name} ===")
                print(df.to_string())
        except Exception as e3:
            print(f"pandas error: {e3}")
